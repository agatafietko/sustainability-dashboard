"""
Fix Excel File - Ensure Headers Are in First Row
This will recreate the Excel file ensuring headers are properly set
"""

import pandas as pd
import openpyxl
from openpyxl import load_workbook

print("="*60)
print("Fixing Excel File Headers")
print("="*60)

# Load existing Excel to check
try:
    wb = load_workbook('Power_BI_Combined_Data.xlsx')
    print("\nCurrent Excel file structure:")
    for sheet_name in wb.sheetnames:
        sheet = wb[sheet_name]
        print(f"\nSheet: {sheet_name}")
        print(f"  First row: {[cell.value for cell in sheet[1]][:5]}...")
except Exception as e:
    print(f"Error reading Excel: {e}")

# Load all CSV files fresh
print("\n" + "="*60)
print("Loading CSV files...")
print("="*60)

sdg_lookup = pd.read_csv('sdg_lookup.csv')
publications = pd.read_csv('Publications_cleaned.csv')
keywords = pd.read_csv('Keywords_cleaned.csv')
sdg_mappings = pd.read_csv('SDG_Mappings_cleaned.csv')

print(f"✓ Loaded all CSV files")

# Create new Excel file with explicit headers
print("\n" + "="*60)
print("Creating fixed Excel file...")
print("="*60)

excel_filename = 'Power_BI_Combined_Data_Fixed.xlsx'

try:
    with pd.ExcelWriter(excel_filename, engine='openpyxl') as writer:
        # Write each dataframe - headers will be in row 1 automatically
        sdg_lookup.to_excel(writer, sheet_name='sdg_lookup', index=False, header=True)
        print(f"✓ Created sheet: sdg_lookup")
        print(f"  Headers: {list(sdg_lookup.columns)}")
        
        publications.to_excel(writer, sheet_name='Publications', index=False, header=True)
        print(f"✓ Created sheet: Publications")
        print(f"  Headers: {list(publications.columns)[:5]}...")
        
        keywords.to_excel(writer, sheet_name='Keywords', index=False, header=True)
        print(f"✓ Created sheet: Keywords")
        print(f"  Headers: {list(keywords.columns)}")
        
        sdg_mappings.to_excel(writer, sheet_name='SDG_Mappings', index=False, header=True)
        print(f"✓ Created sheet: SDG_Mappings")
        print(f"  Headers: {list(sdg_mappings.columns)}")
    
    print(f"\n✅ Successfully created: {excel_filename}")
    
    # Verify headers are there
    wb_new = load_workbook(excel_filename)
    sheet = wb_new['Publications']
    print(f"\nVerification - First row of Publications sheet:")
    print(f"  {[cell.value for cell in sheet[1]][:5]}...")
    
except Exception as e:
    print(f"❌ Error: {e}")
    exit(1)

print("\n" + "="*60)
print("✅ FIXED FILE READY!")
print("="*60)
print(f"\nNew file: {excel_filename}")
print("\nNext steps:")
print("1. Delete the old file from OneDrive")
print("2. Upload this new file: Power_BI_Combined_Data_Fixed.xlsx")
print("3. In Power BI, when importing, make sure 'First row as headers' is checked")
print("4. The column names should now appear correctly!")





