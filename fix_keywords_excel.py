"""
Fix Keywords Sheet in Excel File
Ensure Keywords sheet has proper headers in row 1
"""

import pandas as pd
import openpyxl
from openpyxl import load_workbook

print("="*60)
print("Fixing Keywords Sheet in Excel File")
print("="*60)

# Load Keywords CSV to get correct data
print("\nLoading Keywords CSV...")
keywords_df = pd.read_csv('Keywords_cleaned.csv')
print(f"✓ Loaded Keywords: {len(keywords_df)} rows")
print(f"  Headers: {list(keywords_df.columns)}")
print(f"  First row: {keywords_df.iloc[0].tolist()}")

# Load existing Excel file
print("\n" + "="*60)
print("Checking current Excel file...")
print("="*60)

try:
    wb = load_workbook('Power_BI_Combined_Data_Fixed.xlsx')
    sheet = wb['Keywords']
    
    print(f"\nCurrent Keywords sheet:")
    print(f"  Row 1: {[cell.value for cell in sheet[1][:3]]}")
    print(f"  Row 2: {[cell.value for cell in sheet[2][:3]]}")
    
    # Check if row 1 has headers
    row1_values = [cell.value for cell in sheet[1]]
    if row1_values[0] == 'article_uuid' and row1_values[1] == 'keyword':
        print("\n✅ Headers are already in row 1!")
    else:
        print("\n⚠️  Headers are NOT in row 1 - will fix...")
        
except Exception as e:
    print(f"Error reading Excel: {e}")

# Load all other sheets
print("\n" + "="*60)
print("Loading other sheets...")
print("="*60)

sdg_lookup = pd.read_csv('sdg_lookup.csv')
publications = pd.read_csv('Publications_cleaned.csv')
sdg_mappings = pd.read_csv('SDG_Mappings_cleaned.csv')

print(f"✓ Loaded all CSV files")

# Create NEW Excel file with Keywords properly formatted
print("\n" + "="*60)
print("Creating fixed Excel file with proper Keywords headers...")
print("="*60)

excel_filename = 'Power_BI_Combined_Data_Fixed.xlsx'

try:
    with pd.ExcelWriter(excel_filename, engine='openpyxl') as writer:
        # Write each sheet - ensuring headers are in row 1
        
        # 1. sdg_lookup
        sdg_lookup.to_excel(
            writer, 
            sheet_name='sdg_lookup', 
            index=False,
            header=True,
            startrow=0
        )
        print(f"✓ Created sheet: sdg_lookup")
        
        # 2. Publications
        publications.to_excel(
            writer, 
            sheet_name='Publications', 
            index=False,
            header=True,
            startrow=0
        )
        print(f"✓ Created sheet: Publications")
        
        # 3. Keywords - MAKE SURE HEADERS ARE IN ROW 1
        keywords_df.to_excel(
            writer, 
            sheet_name='Keywords', 
            index=False,  # Don't write row index
            header=True,  # Write column headers
            startrow=0     # Start at row 1 (row 0 in 0-indexed = row 1 in Excel)
        )
        print(f"✓ Created sheet: Keywords")
        print(f"  Headers: {list(keywords_df.columns)}")
        print(f"  First data row: {keywords_df.iloc[0].tolist()}")
        
        # 4. SDG_Mappings
        sdg_mappings.to_excel(
            writer, 
            sheet_name='SDG_Mappings', 
            index=False,
            header=True,
            startrow=0
        )
        print(f"✓ Created sheet: SDG_Mappings")
    
    # Verify the Keywords sheet
    print("\n" + "="*60)
    print("Verifying Keywords sheet...")
    print("="*60)
    
    wb_new = load_workbook(excel_filename)
    sheet = wb_new['Keywords']
    
    # Check row 1 (should be headers)
    row1 = [cell.value for cell in sheet[1]]
    print(f"\nRow 1 (headers): {row1}")
    
    # Check row 2 (should be data)
    if len(list(sheet.rows)) > 1:
        row2 = [cell.value for cell in sheet[2]][:2]
        print(f"Row 2 (first data): {row2}")
    
    # Verify headers are correct
    if row1[0] == 'article_uuid' and row1[1] == 'keyword':
        print("\n✅ SUCCESS! Keywords sheet has proper headers in row 1!")
    else:
        print(f"\n❌ ERROR: Headers not correct. Row 1 = {row1}")
    
    print(f"\n✅ Excel file created: {excel_filename}")
    
except Exception as e:
    print(f"❌ Error: {e}")
    import traceback
    traceback.print_exc()
    exit(1)

print("\n" + "="*60)
print("✅ FIXED FILE READY!")
print("="*60)
print(f"\nFile: {excel_filename}")
print("\nKeywords sheet now has:")
print("  Row 1: article_uuid | keyword (HEADERS)")
print("  Row 2+: [data rows]")
print("\nNext steps:")
print("1. Delete old dataset in Power BI Service")
print("2. Upload this fixed file to OneDrive")
print("3. Create new dataset - Keywords should now have proper headers!")





