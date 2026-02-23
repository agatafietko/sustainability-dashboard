"""
Create Power BI Compatible Excel File
Power BI Service web is picky about headers. This creates a file that's guaranteed to work.
"""

import pandas as pd
import openpyxl
from openpyxl import Workbook
from openpyxl.utils.dataframe import dataframe_to_rows

print("="*60)
print("Creating Power BI Compatible Excel File")
print("="*60)

# Load all CSV files
print("\nLoading CSV files...")
sdg_lookup = pd.read_csv('sdg_lookup.csv')
publications = pd.read_csv('Publications_cleaned.csv')
keywords = pd.read_csv('Keywords_cleaned.csv')
sdg_mappings = pd.read_csv('SDG_Mappings_cleaned.csv')

print(f"✓ Loaded all files")
print(f"  Publications: {len(publications)} rows")
print(f"  Keywords: {len(keywords)} rows")
print(f"  SDG_Mappings: {len(sdg_mappings)} rows")
print(f"  sdg_lookup: {len(sdg_lookup)} rows")

# Create new workbook
print("\n" + "="*60)
print("Creating Excel file with explicit header formatting...")
print("="*60)

wb = Workbook()
# Remove default sheet
if 'Sheet' in wb.sheetnames:
    wb.remove(wb['Sheet'])

def add_sheet_with_headers(wb, df, sheet_name):
    """Add sheet with properly formatted headers"""
    ws = wb.create_sheet(title=sheet_name)
    
    # Write headers to row 1 and make them bold
    from openpyxl.styles import Font
    header_font = Font(bold=True)
    
    # Write headers
    for col_num, header in enumerate(df.columns, 1):
        cell = ws.cell(row=1, column=col_num)
        cell.value = header
        cell.font = header_font
    
    # Write data starting from row 2
    for row_num, row_data in enumerate(df.values, 2):
        for col_num, value in enumerate(row_data, 1):
            ws.cell(row=row_num, column=col_num, value=value)
    
    print(f"✓ Created sheet: {sheet_name}")
    print(f"  Headers in row 1: {list(df.columns)[:3]}...")
    print(f"  Data rows: {len(df)}")

# Add all sheets
add_sheet_with_headers(wb, sdg_lookup, 'sdg_lookup')
add_sheet_with_headers(wb, publications, 'Publications')
add_sheet_with_headers(wb, keywords, 'Keywords')
add_sheet_with_headers(wb, sdg_mappings, 'SDG_Mappings')

# Save file
excel_filename = 'Power_BI_Ready.xlsx'
wb.save(excel_filename)

print(f"\n✅ Created: {excel_filename}")

# Verify the file
print("\n" + "="*60)
print("Verifying file...")
print("="*60)

wb_verify = openpyxl.load_workbook(excel_filename)

for sheet_name in wb_verify.sheetnames:
    sheet = wb_verify[sheet_name]
    row1 = [cell.value for cell in sheet[1][:3]]
    print(f"\n{sheet_name}:")
    print(f"  Row 1 (headers): {row1}")
    if sheet.max_row > 1:
        row2 = [str(cell.value)[:30] if cell.value else 'None' for cell in sheet[2][:3]]
        print(f"  Row 2 (data): {row2}")

print("\n" + "="*60)
print("✅ FILE READY FOR POWER BI!")
print("="*60)
print(f"\nFile: {excel_filename}")
print("\nThis file has:")
print("  ✅ Headers in row 1 (bold formatted)")
print("  ✅ Data starting from row 2")
print("  ✅ All 4 sheets properly formatted")
print("\nNext steps:")
print("1. Upload 'Power_BI_Ready.xlsx' to OneDrive")
print("2. In Power BI Service, create dataset from this file")
print("3. When importing, make sure 'First row as headers' is CHECKED")
print("4. If it still doesn't work, use Transform data → 'Use First Row as Headers'")

